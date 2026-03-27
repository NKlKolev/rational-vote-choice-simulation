from pathlib import Path
import os
import math
import random
import json
import matplotlib.pyplot as plt
import openpyxl
import time  
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet

BASE_DIR = Path(__file__).resolve().parent
FONT_PATH = BASE_DIR / "fonts" / "DejaVuSans.ttf"

FONT_NAME = "Helvetica"

print("DEBUG FONT PATH:", FONT_PATH)
print("DEBUG FONT EXISTS:", FONT_PATH.exists())

if FONT_PATH.exists():
    FONT_NAME = "DejaVuSans"
    pdfmetrics.registerFont(TTFont(FONT_NAME, str(FONT_PATH)))

YES_THRESHOLD = 0.10
NO_THRESHOLD = -0.10
IDEOLOGY_WEIGHT = 0.28
PARTY_WEIGHT = 0.40
SALIENCE_WEIGHT = 0.17
RELATION_WEIGHT = 0.35


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def find_workbook_path():
    data_dir = Path("data")
    preferred = data_dir / "bots.xlsx"
    if preferred.exists():
        return preferred

    matches = sorted(data_dir.glob("*.xlsx"))
    if matches:
        return matches[0]

    return None


def as_float(value, default=0.0):
    if value is None or value == "":
        return default
    return float(value)


def load_bots_from_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bots"]

    headers = [cell.value for cell in ws[1]]
    bots = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        record = dict(zip(headers, row))

        name = record.get("name")
        party = record.get("party")
        if not name or not party:
            continue

        bot = {
            "name": str(name),
            "party": str(party),
            "discipline": as_float(record.get("discipline"), 0.7),
            "rebellion": as_float(record.get("rebellion"), 0.2),
            "emotionality": as_float(record.get("emotionality"), 0.3),
            "opportunism": as_float(record.get("opportunism"), 0.25),
            "corruption_susceptibility": as_float(record.get("corruption_susceptibility"), 0.2),
            "ideology": {
                "economic_left_right": as_float(record.get("economic_left_right")),
                "social_progressive_traditionalist": as_float(record.get("social_progressive_traditionalist")),
                "nationalist_internationalist": as_float(record.get("nationalist_internationalist")),
                "institutionalist_populist": as_float(record.get("institutionalist_populist")),
            },
            "salience": {
                "family_values": as_float(record.get("family_values")),
                "civil_rights": as_float(record.get("civil_rights")),
                "economy": as_float(record.get("economy")),
                "security": as_float(record.get("security")),
                "infrastructure": as_float(record.get("infrastructure")),
                "public_spending": as_float(record.get("public_spending")),
                "regional_development": as_float(record.get("regional_development")),
                "controversy": as_float(record.get("controversy"), 0.5),
                "public_support": as_float(record.get("public_support"), 0.5),
                "fiscal_impact": as_float(record.get("fiscal_impact"), 0.5),
                "international_alignment": as_float(record.get("international_alignment"), 0.5),
                "urgency": as_float(record.get("urgency"), 0.5),
            }
        }
        bots.append(bot)

    return bots


def clamp(value, min_value=-1.0, max_value=1.0):
    return max(min(value, max_value), min_value)

def normalize(text):
    return text.strip().lower()


def canonical_party_name(name):
    normalized = normalize(str(name))

    aliases = {
        "traditional alliance": "ПКП-НН",
        "liberal democrats": "Демократично Пустиняково",
        "centrist union": "ДПС",
        "regional development movement": "ДПС",
        "national conservatives": "ГЕРП",
        "civic liberal party": "Демократично Пустиняково",
        "social solidarity party": "ПКП-НН",
        "green future alliance": "СХДС",
        "people’s voice front": "ГЕРП",
        "people's voice front": "ГЕРП"
    }

    return aliases.get(normalized, str(name))


PARTY_RELATIONS = {
    "ГЕРП": {
        "ПКП-НН": 0.50,
        "ДПС": -0.60,
        "Демократично Пустиняково": -0.50,
        "СХДС": 0.20,
    },
    "ПКП-НН": {
        "ГЕРП": 0.50,
        "ДПС": -0.85,
        "Демократично Пустиняково": -0.90,
        "СХДС": -0.15,
    },
    "ДПС": {
        "ПКП-НН": -0.95,
        "ГЕРП": -0.80,
        "Демократично Пустиняково": 0.45,
        "СХДС": -0.95,
    },
    "Демократично Пустиняково": {
        "ПКП-НН": -0.95,
        "ДПС": 0.50,
        "ГЕРП": -0.75,
        "СХДС": 0.15,
    },
    "СХДС": {
        "ПКП-НН": 0.30,
        "ДПС": -0.95,
        "ГЕРП": 0.20,
        "Демократично Пустиняково": 0.15,
    },
}
PARTY_PROFILES = {
    "ПКП-НН": {
        "policy_profile": {
            "social_progressive_traditionalist": 0.9,
            "civil_rights": -0.8,
            "economic_left_right": -0.4,
            "security": 0.95,
            "public_spending": 0.5,
            "regional_development": 0.6,
            "international_alignment": -0.6,
        },
        "salience_profile": {
            "family_values": 0.8,
            "civil_rights": 0.6,
            "economy": 0.7,
            "security": 1.0,
            "infrastructure": 0.6,
            "public_spending": 0.7,
            "regional_development": 0.8,
            "controversy": 0.8,
            "public_support": 0.0,
            "fiscal_impact": 0.7,
            "international_alignment": -0.5,
            "urgency": 0.8,
        },
        "strategy_profile": {
            "public_support_sensitivity": 0.4,
            "controversy_aversion": 0.2,
            "urgency_responsiveness": 0.7,
            "fiscal_sensitivity": 0.7,
            "international_alignment_sensitivity": 0.2,
            "compromise_willingness": 0.3,
            "opposition_instinct": 0.6,
            "governing_instinct": 1.0,
        },
    },

    "ГЕРП": {
        "policy_profile": {
            "social_progressive_traditionalist": 0.2,
            "civil_rights": 0.1,
            "economic_left_right": 0.4,
            "security": 0.6,
            "public_spending": 0.2,
            "regional_development": 0.7,
            "international_alignment": 0.6,
        },
        "salience_profile": {
            "family_values": 0.4,
            "civil_rights": 0.4,
            "economy": 0.9,
            "security": 0.8,
            "infrastructure": 0.9,
            "public_spending": 0.7,
            "regional_development": 1.0,
            "controversy": 0.6,
            "public_support": 0.9,
            "fiscal_impact": 0.8,
            "international_alignment": 0.7,
            "urgency": 0.7,
        },
        "strategy_profile": {
            "public_support_sensitivity": 0.9,
            "controversy_aversion": 0.7,
            "urgency_responsiveness": 0.8,
            "fiscal_sensitivity": 0.9,
            "international_alignment_sensitivity": 0.7,
            "compromise_willingness": 0.7,
            "opposition_instinct": 0.3,
            "governing_instinct": 0.9,
        },
    },
    "Демократично Пустиняково": {
        "policy_profile": {
            "social_progressive_traditionalist": -0.7,
            "civil_rights": 0.9,
            "economic_left_right": -0.1,
            "security": 0.3,
            "public_spending": 0.4,
            "regional_development": 0.5,
            "international_alignment": 0.9,
        },
        "salience_profile": {
            "family_values": 0.3,
            "civil_rights": 1.0,
            "economy": 0.7,
            "security": 0.5,
            "infrastructure": 0.6,
            "public_spending": 0.6,
            "regional_development": 0.6,
            "controversy": 0.5,
            "public_support": 0.8,
            "fiscal_impact": 0.6,
            "international_alignment": 1.0,
            "urgency": 0.6,
        },
        "strategy_profile": {
            "public_support_sensitivity": 0.8,
            "controversy_aversion": 0.6,
            "urgency_responsiveness": 0.6,
            "fiscal_sensitivity": 0.6,
            "international_alignment_sensitivity": 1.0,
            "compromise_willingness": 0.6,
            "opposition_instinct": 0.7,
            "governing_instinct": 0.5,
        },
    },

    "ДПС": {
        "policy_profile": {
            "social_progressive_traditionalist": 0.0,
            "civil_rights": 0.3,
            "economic_left_right": 0.1,
            "security": 0.4,
            "public_spending": 0.6,
            "regional_development": 0.8,
            "international_alignment": 0.6,
        },
        "salience_profile": {
            "family_values": 0.4,
            "civil_rights": 0.6,
            "economy": 0.8,
            "security": 0.6,
            "infrastructure": 0.8,
            "public_spending": 0.9,
            "regional_development": 1.0,
            "controversy": 0.9,
            "public_support": 0.6,
            "fiscal_impact": 0.9,
            "international_alignment": 0.7,
            "urgency": 0.7,
        },
        "strategy_profile": {
            "public_support_sensitivity": 0.6,
            "controversy_aversion": 0.4,
            "urgency_responsiveness": 0.7,
            "fiscal_sensitivity": 0.9,
            "international_alignment_sensitivity": 0.7,
            "compromise_willingness": 0.9,
            "opposition_instinct": 0.5,
            "governing_instinct": 0.8,
        },
    },
    "СХДС": {
        "policy_profile": {
            "social_progressive_traditionalist": -0.6,
            "civil_rights": 0.8,
            "economic_left_right": -0.3,
            "security": 0.2,
            "public_spending": 0.6,
            "regional_development": 0.6,
            "international_alignment": 0.8,
        },
        "salience_profile": {
            "family_values": 0.3,
            "civil_rights": 0.9,
            "economy": 0.6,
            "security": 0.4,
            "infrastructure": 0.7,
            "public_spending": 0.7,
            "regional_development": 0.7,
            "controversy": 0.7,
            "public_support": 0.7,
            "fiscal_impact": 0.6,
            "international_alignment": 0.9,
            "urgency": 0.8,
        },
        "strategy_profile": {
            "public_support_sensitivity": 0.7,
            "controversy_aversion": 0.5,
            "urgency_responsiveness": 0.8,
            "fiscal_sensitivity": 0.6,
            "international_alignment_sensitivity": 0.9,
            "compromise_willingness": 0.6,
            "opposition_instinct": 0.7,
            "governing_instinct": 0.5,
        },
    },
}

def get_party_relation(from_party, to_party):
    from_party = canonical_party_name(from_party)
    to_party = canonical_party_name(to_party)
    return PARTY_RELATIONS.get(from_party, {}).get(to_party, 0.0)

def calculate_single_party_position(party_name, proposal):
    profile = PARTY_PROFILES.get(party_name)
    if not profile:
        return 0.0

    policy_profile = profile["policy_profile"]
    salience_profile = profile["salience_profile"]
    strategy_profile = profile["strategy_profile"]
    effects = proposal.get("effects", {})

    stance = 0.0

    # 1. Core ideological / policy alignment
    policy_dimensions = [
        "social_progressive_traditionalist",
        "civil_rights",
        "economic_left_right",
        "security",
        "public_spending",
        "regional_development",
        "international_alignment",
    ]

    for dim in policy_dimensions:
        proposal_value = effects.get(dim, 0.0)
        party_value = policy_profile.get(dim, 0.0)
        stance += proposal_value * party_value * 0.35

    # 2. Issue salience effects
    salience_mapping = {
        "civil_rights": "civil_rights",
        "economy": "economy",
        "security": "security",
        "infrastructure": "infrastructure",
        "public_spending": "public_spending",
        "regional_development": "regional_development",
        "controversy": "controversy",
        "public_support": "public_support",
        "fiscal_impact": "fiscal_impact",
        "international_alignment": "international_alignment",
        "urgency": "urgency",
    }

    for effect_key, salience_key in salience_mapping.items():
        proposal_value = effects.get(effect_key, 0.0)
        salience_value = salience_profile.get(salience_key, 0.0)
        stance += proposal_value * salience_value * 0.15

    # 3. Strategic response modifiers
    stance += effects.get("public_support", 0.0) * strategy_profile.get("public_support_sensitivity", 0.0) * 0.20
    stance -= effects.get("controversy", 0.0) * strategy_profile.get("controversy_aversion", 0.0) * 0.20
    stance += effects.get("urgency", 0.0) * strategy_profile.get("urgency_responsiveness", 0.0) * 0.15
    stance += effects.get("fiscal_impact", 0.0) * strategy_profile.get("fiscal_sensitivity", 0.0) * 0.15
    stance += effects.get("international_alignment", 0.0) * strategy_profile.get("international_alignment_sensitivity", 0.0) * 0.15

    # 4. Proposer / coalition relation effect
    proposed_by_party = proposal.get("proposed_by_party")
    relation_score = 0.0

    if proposed_by_party:
        relation_score = get_party_relation(party_name, proposed_by_party)

        # Base relation effect
        stance += relation_score * 0.25

        # Coalition reinforcement:
        # strong allies/supporters move a bit more toward support,
        # strong opponents move a bit more toward opposition
        if relation_score >= 0.4:
            stance += relation_score * 0.10
        elif relation_score <= -0.4:
            stance += relation_score * 0.10

    # 5. Governing/opposition instincts
    stance += strategy_profile.get("governing_instinct", 0.0) * effects.get("urgency", 0.0) * 0.05
    stance -= strategy_profile.get("opposition_instinct", 0.0) * effects.get("controversy", 0.0) * 0.05

    # 6. Compromise dynamics
    compromise = strategy_profile.get("compromise_willingness", 0.0)
    uncertainty = 1.0 - min(abs(stance), 1.0)

    # Pull uncertain positions toward the center
    stance += (-stance) * compromise * uncertainty * 0.40

    return clamp(stance, -1.0, 1.0)
def calculate_party_positions(proposal):
    positions = {}

    for party_name in PARTY_PROFILES.keys():
        positions[party_name] = round(
            calculate_single_party_position(party_name, proposal),
            2
        )

    return positions

# Helper to get unique party names in order of appearance
def get_party_names(bots):
    seen = []
    for bot in bots:
        party = bot.get("party", "")
        if party and party not in seen:
            seen.append(party)
    return seen


def score_bot(bot, proposal):
    proposal_type = proposal.get("type", "policy")
    ideology_score = 0.0
    salience_score = 0.0

    effects = proposal.get("effects", {})
    ideology = bot.get("ideology", {})
    salience = bot.get("salience", {})

    if proposal_type == "pm_election":
        party = canonical_party_name(bot.get("party", ""))
        candidate_party = canonical_party_name(proposal.get("candidate_party", ""))
        relation_score = get_party_relation(party, candidate_party)
        discipline = bot.get("discipline", 0.5)
        rebellion = bot.get("rebellion", 0.2)

        party_positions = proposal.get("party_positions", {})
        normalized_party_positions = {
            normalize(str(key)): value for key, value in party_positions.items()
        }
        party_pressure = normalized_party_positions.get(normalize(str(party)), 0.0)

        yes_threshold = YES_THRESHOLD
        no_threshold = NO_THRESHOLD

        #Hard party-line override for PM election
        if party_pressure <= -0.95:
            loyalty_chance = 0.55 + discipline * 0.2

            if random.random() < loyalty_chance:
                vote = "NO"
                reason = "партията силно се противопоставя на кандидата"
            else:
                vote = "ABSTAIN" if random.random() < 0.8 else "YES"
                reason = "вътрешно отклонение от партийната линия"

            return {
                "name": bot["name"],
                "party": bot["party"],
                "score": -1.0,
                "vote": vote,
                "reason": reason,
                "party_pressure": round(party_pressure, 3),
                "ideology_score": 0.0,
                "salience_score": 0.0,
                "relation_score": round(relation_score, 3),
                "randomness": 0.0,
                "yes_threshold": round(yes_threshold, 3),
                "no_threshold": round(no_threshold, 3)
            }

        if party_pressure <= -0.9 and discipline >= 0.65:
            vote = "NO"
            reason = "партията твърдо се противопоставя на кандидата за министър-председател"
            return {
                "name": bot["name"],
                "party": bot["party"],
                "score": -1.0,
                "vote": vote,
                "reason": reason,
                "party_pressure": round(party_pressure, 3),
                "ideology_score": 0.0,
                "salience_score": 0.0,
                "relation_score": round(relation_score, 3),
                "randomness": 0.0,
                "yes_threshold": round(yes_threshold, 3),
                "no_threshold": round(no_threshold, 3)
            }
        if normalize(str(party)) == normalize(str(candidate_party)):
            base_score = 0.98
            random_scale = 0.02 + rebellion * 0.05
            randomness = random.uniform(-random_scale, random_scale) * (1.0 - discipline * 0.3)
            score = clamp(base_score + randomness)

            if score >= yes_threshold:
                vote = "YES"
                reason = "подкрепя кандидата за министър-председател на собствената си партия"
            else:
                vote = "ABSTAIN"
                reason = "собствена партийна кандидатура, но има вътрешно колебание"
        else:
            base_score = relation_score * (0.75 + discipline * 0.25)

            random_scale = 0.10 + rebellion * 0.25
            if abs(relation_score) <= 0.25:
                random_scale += 0.08
            randomness = random.uniform(-random_scale, random_scale)
            score = clamp(base_score + randomness)

            if abs(relation_score) >= 0.60:
                yes_threshold -= 0.01
                no_threshold += 0.01

            if score >= yes_threshold:
                vote = "YES"
                reason = "подкрепя кандидата заради отношенията между партиите"
            elif score <= no_threshold:
                vote = "NO"
                reason = "се противопоставя на кандидата заради отношенията между партиите"
            else:
                vote = "ABSTAIN"
                reason = "отношенията между партиите и вътрешната динамика не решават еднозначно този вот"

        return {
            "name": bot["name"],
            "party": bot["party"],
            "score": round(score, 3),
            "vote": vote,
            "reason": reason,
            "party_pressure": round(party_pressure, 3),
            "ideology_score": 0.0,
            "salience_score": 0.0,
            "relation_score": round(relation_score, 3),
            "randomness": round(randomness, 3),
            "yes_threshold": round(yes_threshold, 3),
            "no_threshold": round(no_threshold, 3)
        }

    if proposal_type != "policy":
        return {
            "name": bot["name"],
            "party": bot["party"],
            "score": 0.0,
            "vote": "ABSTAIN",
            "reason": f"неподдържан тип предложение: {proposal_type}",
            "party_pressure": 0.0,
            "ideology_score": 0.0,
            "salience_score": 0.0,
            "relation_score": 0.0,
            "randomness": 0.0,
            "yes_threshold": 0.0,
            "no_threshold": 0.0
        }

    if "social_progressive_traditionalist" in effects:
        proposal_pos = effects["social_progressive_traditionalist"]
        bot_pos = ideology.get("social_progressive_traditionalist", 0.0)
        ideology_score += proposal_pos * bot_pos

    if "economic_left_right" in effects:
        proposal_pos = effects["economic_left_right"]
        bot_pos = ideology.get("economic_left_right", 0.0)
        ideology_score += proposal_pos * bot_pos

    party = canonical_party_name(bot.get("party", ""))

    # AUTO-CALCULATED PARTY POSITIONS FOR POLICY VOTING
    party_positions = calculate_party_positions(proposal)
    party_pressure = party_positions.get(party, 0.0)

    proposed_by_party = proposal.get("proposed_by_party")
    relation_score = 0.0
    coalition_bonus = 0.0

    if proposed_by_party:
        relation_score = get_party_relation(party, proposed_by_party)

        # Coalition dynamics at MP level
        if relation_score >= 0.4:
            coalition_bonus = relation_score * 0.12
        elif relation_score <= -0.4:
            coalition_bonus = relation_score * 0.12

    if "civil_rights" in effects:
        civil_rights_salience = salience.get("civil_rights", 0.0)
        salience_score += civil_rights_salience * effects["civil_rights"] * 0.3

    if "family_values_conflict" in effects:
        family_values_salience = salience.get("family_values", 0.0)
        salience_score -= family_values_salience * effects["family_values_conflict"] * 0.3

    if "economy" in effects:
        economy_salience = salience.get("economy", 0.0)
        salience_score += economy_salience * effects["economy"] * 0.3

    if "security" in effects:
        security_salience = salience.get("security", 0.0)
        salience_score += security_salience * effects["security"] * 0.3

    if "infrastructure" in effects:
        infrastructure_salience = salience.get("infrastructure", 0.0)
        salience_score += infrastructure_salience * effects["infrastructure"] * 0.3

    if "public_spending" in effects:
        spending_salience = salience.get("public_spending", 0.0)
        salience_score += spending_salience * effects["public_spending"] * 0.3

    if "regional_development" in effects:
        regional_salience = salience.get("regional_development", 0.0)
        salience_score += regional_salience * effects["regional_development"] * 0.3

    salience_score += effects.get("controversy", 0.0) * salience.get("controversy", 0.0) * 0.20
    salience_score += effects.get("public_support", 0.0) * salience.get("public_support", 0.0) * 0.20
    salience_score += effects.get("fiscal_impact", 0.0) * salience.get("fiscal_impact", 0.0) * 0.20
    salience_score += effects.get("international_alignment", 0.0) * salience.get("international_alignment", 0.0) * 0.20
    salience_score += effects.get("urgency", 0.0) * salience.get("urgency", 0.0) * 0.20

    discipline = bot.get("discipline", 0.5)
    rebellion = bot.get("rebellion", 0.2)
    controversy = abs(effects.get("controversy", 0.0))

    relation_alignment = relation_score * (0.7 + discipline * 0.3)
    party_line_effect = party_pressure * (0.55 + discipline * 0.45)

    # Increased randomness baseline and scaling
    random_scale = 0.18 + controversy * 0.15

    # More randomness when party signals are weak
    if abs(party_pressure) <= 0.2:
        random_scale += 0.06

    # Additional randomness when relations are unclear
    if abs(relation_score) <= 0.2:
        random_scale += 0.05

    # Stronger stochastic behavior (non-linear)
    randomness = random.uniform(-1, 1) * (rebellion * random_scale * 1.5)

    # Slightly reduce deterministic dominance, increase randomness influence
    total_score = (
            ideology_score * (IDEOLOGY_WEIGHT * 0.9)
            + party_line_effect * (PARTY_WEIGHT * 0.9)
            + salience_score * (SALIENCE_WEIGHT * 0.9)
            + relation_alignment * (RELATION_WEIGHT * 0.9)
            + coalition_bonus
            + randomness
        )
    )

    # --- NON-RATIONAL LAYERS ---

    # 1. Emotional voting (irrational swings)
    emotionality = bot.get("emotionality", 0.3)
    emotion = random.uniform(-1, 1) * emotionality * 0.4
    total_score += emotion

    # 2. Opportunistic flips (strategic hypocrisy)
    opportunism = bot.get("opportunism", 0.25)
    if random.random() < opportunism * 0.3:
        total_score *= -0.5

    # 3. Corruption / material incentives
    corruption = bot.get("corruption_susceptibility", 0.2)
    corruption_push = effects.get("fiscal_impact", 0.0) * corruption * 0.3
    total_score += corruption_push

    total_score = clamp(total_score)
    # --- COMPROMISE EFFECT ---
    compromise = bot.get("opportunism", 0.25) * 0.0  # placeholder, not used directly
    party_profile = PARTY_PROFILES.get(party, {})
    strategy_profile = party_profile.get("strategy_profile", {})
    compromise_willingness = strategy_profile.get("compromise_willingness", 0.0)

    score_uncertainty = 1.0 - min(abs(total_score), 1.0)

    # Pull borderline votes toward the center for compromise-prone parties
    total_score += (-total_score) * compromise_willingness * score_uncertainty * 0.30
    total_score = clamp(total_score)
    # --- PARTY FRACTURE EFFECT ---

    faction_noise = random.uniform(-1, 1) * bot.get("rebellion", 0.2) * 0.35

    # stronger fractures when party pressure is high (rebels react)
    if abs(party_pressure) > 0.6:
        faction_noise *= 1.5
    # Rare rebellion shock
    if random.random() < bot.get("rebellion", 0.2) * 0.08:
        total_score *= -1

    total_score += faction_noise
    yes_threshold = YES_THRESHOLD + (controversy * 0.04)
    no_threshold = NO_THRESHOLD - (controversy * 0.04)

    # Make thresholds less rigid → more swing outcomes
    if abs(party_pressure) >= 0.75:
        yes_threshold -= 0.02
        no_threshold += 0.02

    if abs(relation_score) >= 0.60:
        yes_threshold -= 0.01
        no_threshold += 0.01
        
    if party_pressure <= -0.9 and discipline >= 0.65:
        vote = "NO"
        reasons = ["партията твърдо се противопоставя на предложението"]
        return {
            "name": bot["name"],
            "party": bot["party"],
            "vote": vote,
            "score": round(total_score, 3),
            "reason": "; ".join(reasons),
            "party_pressure": round(party_pressure, 3),
            "ideology_score": round(ideology_score, 3),
            "salience_score": round(salience_score, 3),
            "relation_score": round(relation_score, 3),
            "randomness": round(randomness, 3),
            "coalition_bonus": round(coalition_bonus, 3),
            "yes_threshold": round(yes_threshold, 3),
            "no_threshold": round(no_threshold, 3)
            }
    if total_score >= yes_threshold:
        vote = "YES"
    elif total_score <= no_threshold:
        vote = "NO"
    else:
        vote = "ABSTAIN"

    reasons = []

    if ideology_score > 0.2:
        reasons.append("идеологията подкрепя предложението")
    elif ideology_score < -0.2:
        reasons.append("идеологията се противопоставя на предложението")

    if party_pressure > 0.3:
        reasons.append("партията подкрепя предложението")
    elif party_pressure < -0.3:
        reasons.append("партията се противопоставя на предложението")

    if relation_score > 0.3:
        reasons.append("предложението е внесено от съюзническа партия")
    elif relation_score < -0.3:
        reasons.append("предложението е внесено от противникова партия")

    if abs(effects.get("controversy", 0.0)) >= 0.5:
        reasons.append("спорен законопроект")

    if abs(effects.get("fiscal_impact", 0.0)) >= 0.5:
        reasons.append("силен бюджетен ефект")

    if abs(effects.get("international_alignment", 0.0)) >= 0.5:
        reasons.append("има силно международно измерение")

    if effects.get("urgency", 0.0) >= 0.5:
        reasons.append("висока спешност")

    if salience_score > 0.5:
        reasons.append("важен въпрос за този народен представител")
    elif salience_score < -0.5:
        reasons.append("важен конфликтен въпрос за този народен представител")

    if not reasons:
        reasons = ["смесени или слаби сигнали"]

    return {
        "name": bot["name"],
        "party": bot["party"],
        "vote": vote,
        "score": round(total_score, 3),
        "reason": "; ".join(reasons),

        "party_pressure": round(party_pressure, 3),
        "ideology_score": round(ideology_score, 3),
        "salience_score": round(salience_score, 3),
        "relation_score": round(relation_score, 3),
        "randomness": round(randomness, 3),
        "coalition_bonus": round(coalition_bonus, 3),

        "yes_threshold": round(yes_threshold, 3),
        "no_threshold": round(no_threshold, 3)
    }


def run_vote(bots, proposal):
    results = []

    current_yes = 0
    current_no = 0
    current_abstain = 0

    for bot in bots:
        result = score_bot(bot, proposal)

        # --- BANDWAGON EFFECT ---
        total_votes = current_yes + current_no + current_abstain

        if total_votes > 5:  # only after some votes exist
            yes_ratio = current_yes / total_votes
            no_ratio = current_no / total_votes

            bandwagon_strength = 0.25 + bot.get("opportunism", 0.2)

            if yes_ratio > 0.6 and abs(result["score"]) < 0.5:
                if random.random() < bandwagon_strength:
                    result["vote"] = "YES"
                    result["reason"] += "; повлияно от мнозинството"


            elif no_ratio > 0.6 and abs(result["score"]) < 0.5:
                if random.random() < bandwagon_strength:
                    result["vote"] = "NO"
                    result["reason"] += "; повлияно от мнозинството"

        # update counters AFTER influence
        if result["vote"] == "YES":
            current_yes += 1
        elif result["vote"] == "NO":
            current_no += 1
        else:
            current_abstain += 1

        results.append(result)

    return results
def get_representative_vote(bots, proposal, n_runs=100, pass_threshold=33, progress_callback=None):
    simulations = []

    for i in range(n_runs):
        results = run_vote(bots, proposal)
        totals = count_votes(results)
        party_totals = count_votes_by_party(results)
        bill_passed = totals["YES"] >= pass_threshold

        simulations.append({
            "results": results,
            "totals": totals,
            "party_totals": party_totals,
            "bill_passed": bill_passed,
        })

        if progress_callback:
            displayed_step = min(65, max(1, int(((i + 1) / n_runs) * 65)))
            progress_callback(displayed_step, 65)
            time.sleep(0.03)

    # Split simulations into PASS / FAIL groups
    passed_runs = [sim for sim in simulations if sim["bill_passed"]]
    failed_runs = [sim for sim in simulations if not sim["bill_passed"]]

    # Choose the dominant class
    if len(passed_runs) >= len(failed_runs):
        dominant_runs = passed_runs if passed_runs else simulations
        dominant_outcome = "PASS"
    else:
        dominant_runs = failed_runs if failed_runs else simulations
        dominant_outcome = "FAIL"

    # Compute average totals within the dominant class
    avg_yes = sum(sim["totals"]["YES"] for sim in dominant_runs) / len(dominant_runs)
    avg_no = sum(sim["totals"]["NO"] for sim in dominant_runs) / len(dominant_runs)
    avg_abstain = sum(sim["totals"]["ABSTAIN"] for sim in dominant_runs) / len(dominant_runs)

    # Find the single run closest to those averages
    def distance(sim):
        return (
            abs(sim["totals"]["YES"] - avg_yes)
            + abs(sim["totals"]["NO"] - avg_no)
            + abs(sim["totals"]["ABSTAIN"] - avg_abstain)
        )

    representative = min(dominant_runs, key=distance)

    # Average party totals within dominant class
    average_party_totals = {}
    all_parties = set()
    for sim in dominant_runs:
        all_parties.update(sim["party_totals"].keys())

    for party in all_parties:
        yes_avg = sum(sim["party_totals"].get(party, {}).get("YES", 0) for sim in dominant_runs) / len(dominant_runs)
        no_avg = sum(sim["party_totals"].get(party, {}).get("NO", 0) for sim in dominant_runs) / len(dominant_runs)
        abstain_avg = sum(sim["party_totals"].get(party, {}).get("ABSTAIN", 0) for sim in dominant_runs) / len(dominant_runs)

        average_party_totals[party] = {
            "YES": round(yes_avg, 2),
            "NO": round(no_avg, 2),
            "ABSTAIN": round(abstain_avg, 2),
        }

    return {
        "results": representative["results"],
        "totals": representative["totals"],
        "party_totals": representative["party_totals"],
        "bill_passed": representative["bill_passed"],
        "meta": {
            "n_runs": n_runs,
            "pass_probability": round(len(passed_runs) / n_runs, 3),
            "fail_probability": round(len(failed_runs) / n_runs, 3),
            "dominant_outcome": dominant_outcome,
            "dominant_runs": len(dominant_runs),
            "average_totals_in_dominant_class": {
                "YES": round(avg_yes, 2),
                "NO": round(avg_no, 2),
                "ABSTAIN": round(avg_abstain, 2),
            },
            "average_party_totals_in_dominant_class": average_party_totals,
        }
    }
def generate_pdf_report(proposal, results, totals, party_totals, bill_passed):
    doc = SimpleDocTemplate("parliament_report.pdf")
    styles = getSampleStyleSheet()
    styles["Title"].fontName = FONT_NAME
    styles["Heading2"].fontName = FONT_NAME
    styles["Heading3"].fontName = FONT_NAME
    styles["Normal"].fontName = FONT_NAME


    if FONT_NAME == "Helvetica":
        content = [
            Paragraph(
                "<b>PDF предупреждение:</b> Не е намерен шрифт с поддръжка на кирилица. "
                "Инсталирай DejaVu Sans или Arial Unicode, за да се вижда българският текст правилно.",
                styles["Normal"]
            ),
            Spacer(1, 12)
        ]
    else:
        content = []

    content.append(Paragraph("<b>Конгресен Доклад</b>", styles["Title"]))
    content.append(Spacer(1, 12))

    content.append(Paragraph(f"<b>Предложение:</b> {proposal['title']}", styles["Heading2"]))
    content.append(Paragraph(f"<b>ID на законопроекта:</b> {proposal.get('bill_id', 'Няма ID')}", styles["Normal"]))
    content.append(Spacer(1, 10))

    description = proposal.get("description") or "Няма добавено описание на законопроекта."
    changes_text = proposal.get("changes") or proposal.get("changes_summary") or "Няма добавено описание на промените."

    content.append(Paragraph(f"<b>За какво е законопроектът:</b> {description}", styles["Normal"]))
    content.append(Spacer(1, 8))
    content.append(Paragraph(f"<b>Какво променя:</b> {changes_text}", styles["Normal"]))
    content.append(Spacer(1, 10))

    content.append(Paragraph("<b>Характеристики на политиката:</b>", styles["Heading3"]))
    for key, value in proposal.get("effects", {}).items():
        content.append(Paragraph(f"{key}: {value}", styles["Normal"]))

    content.append(Spacer(1, 10))
    content.append(Paragraph("<b>Допълнителни характеристики:</b>", styles["Heading3"]))
    content.append(Paragraph(f"Спорност: {proposal.get('effects', {}).get('controversy', 0)}", styles["Normal"]))
    content.append(Paragraph(f"Обществена подкрепа: {proposal.get('effects', {}).get('public_support', 0)}", styles["Normal"]))
    content.append(Paragraph(f"Бюджетен ефект: {proposal.get('effects', {}).get('fiscal_impact', 0)}", styles["Normal"]))
    content.append(Paragraph(f"Международно измерение: {proposal.get('effects', {}).get('international_alignment', 0)}", styles["Normal"]))
    content.append(Paragraph(f"Спешност: {proposal.get('effects', {}).get('urgency', 0)}", styles["Normal"]))
    content.append(Spacer(1, 10))
    #content.append(Paragraph("<b>Позиции на партиите:</b>", styles["Heading3"]))  #this puts the party positions within the report, they are temporarily removed!
    #for party, value in proposal.get("party_positions", {}).items():
        #content.append(Paragraph(f"{party}: {value}", styles["Normal"]))

    if proposal.get("proposed_by_party"):
        content.append(Spacer(1, 10))
        content.append(Paragraph(f"<b>Вносител:</b> {proposal['proposed_by_party']}", styles["Normal"]))

    content.append(Spacer(1, 12))

    content.append(Paragraph("<b>Резултати:</b>", styles["Heading3"]))
    content.append(Spacer(1, 8))

    for result in results:
        vote_label = {
            "YES": "ЗА",
            "NO": "ПРОТИВ",
            "ABSTAIN": "ВЪЗДЪРЖАЛ СЕ"
        }.get(result["vote"], result["vote"])
        text = (
            f"{result['name']} ({result['party']}) → {vote_label} | "
            f"резултат={result['score']} | обяснение={result['reason']}"
        )
        content.append(Paragraph(text, styles["Normal"]))
        content.append(Spacer(1, 6))

    content.append(Spacer(1, 12))

    content.append(Paragraph("<b>Общо:</b>", styles["Heading3"]))
    content.append(Paragraph(f"ЗА: {totals['YES']}", styles["Normal"]))
    content.append(Paragraph(f"ПРОТИВ: {totals['NO']}", styles["Normal"]))
    content.append(Paragraph(f"ВЪЗДЪРЖАЛ СЕ: {totals['ABSTAIN']}", styles["Normal"]))

    content.append(Spacer(1, 12))
    content.append(Paragraph("<b>Гласуване по партии:</b>", styles["Heading3"]))

    for party, counts in party_totals.items():
        content.append(
            Paragraph(
                f"{party}: ЗА={counts['YES']}, ПРОТИВ={counts['NO']}, ВЪЗДЪРЖАЛ СЕ={counts['ABSTAIN']}",
                styles["Normal"]
            )
        )

    content.append(Spacer(1, 12))
    status_text = "ПРИЕТО" if bill_passed else "НЕ Е ПРИЕТО"
    content.append(Paragraph(f"<b>Краен резултат:</b> {status_text}", styles["Heading3"]))

    doc.build(content)


def count_votes(results):
    totals = {"YES": 0, "NO": 0, "ABSTAIN": 0}

    for result in results:
        vote = result["vote"]
        totals[vote] += 1

    return totals


# Party-level vote counting
def count_votes_by_party(results):
    party_totals = {}

    for result in results:
        party = result["party"]
        if party not in party_totals:
            party_totals[party] = {"YES": 0, "NO": 0, "ABSTAIN": 0}

        party_totals[party][result["vote"]] += 1

    return party_totals

if __name__ == "__main__":
    workbook_path = find_workbook_path()
    if workbook_path:
        bots = load_bots_from_excel(workbook_path)
        print(f"Заредени народни представители от Excel: {workbook_path}")
    else:
        bots = load_json("data/bots.json")
        print("Заредени народни представители от JSON: data/bots.json")
    proposals = load_json("data/proposals.json")
    party_names = get_party_names(bots)

    add_custom = input("Искаш ли да добавиш ново предложение? (y/n): ").strip().lower()

    if add_custom == "y":
        custom_title = input("Въведи заглавие на предложението: ").strip()
        custom_bill_id = input("Въведи уникален номер на законопроекта: ").strip()
        custom_description = input("Опиши накратко за какво е законопроектът: ").strip()
        custom_changes = input("Опиши накратко какво променя законопроектът: ").strip()

        proposal_type = input("Въведи тип предложение (policy/pm_election): ").strip().lower()
        candidate_party = None
        proposed_by_party = input("Въведи вносител на предложението (по желание): ").strip() or None
        if proposal_type == "pm_election":
            candidate_party = input("Въведи партията на кандидата: ").strip()
        effects = {
            "social_progressive_traditionalist": 0.0,
            "civil_rights": 0.0,
            "family_values_conflict": 0.0,
            "economic_left_right": 0.0,
            "economy": 0.0,
            "security": 0.0,
            "infrastructure": 0.0,
            "public_spending": 0.0,
            "regional_development": 0.0,
            "controversy": 0.0,
            "public_support": 0.0,
            "fiscal_impact": 0.0,
            "international_alignment": 0.0,
            "urgency": 0.0
        }

        party_positions = {party_name: 0.0 for party_name in party_names}

        if proposal_type == "policy":
            print("\nВъведи стойности за ефектите на политиката между -1.0 и 1.0")
            effects["social_progressive_traditionalist"] = float(
                input("Тласка ли тази политика обществото към прогресивни (-1) или традиционни (+1) норми? ").strip() or 0)

            effects["civil_rights"] = float(
                input("Разширява ли (+1) или ограничава (-1) тази политика гражданските права? ").strip() or 0)

            effects["family_values_conflict"] = float(
                input("Противоречи ли (+1) или укрепва (-1) тази политика традиционните семейни ценности? ").strip() or 0)

            effects["economic_left_right"] = float(
                input("Икономически дясна (+1, пазарна) или лява (-1, държавни разходи) ли е тази политика? ").strip() or 0)

            effects["economy"] = float(
                input("Колко икономически важна е тази политика? (0 до 1) ").strip() or 0)

            effects["security"] = float(
                input("Увеличава (+1) или намалява (-1) тази политика сигурността/контрола? ").strip() or 0)

            effects["infrastructure"] = float(
                input("Инфраструктурна политика ли е това? (0 до 1) ").strip() or 0)

            effects["public_spending"] = float(
                input("Включва ли това публични разходи (+1) или съкращения (-1)? ").strip() or 0)

            effects["regional_development"] = float(
                input("Подкрепя ли това регионалното развитие? (0 до 1) ").strip() or 0)
            effects["controversy"] = float(
                input("Колко спорен е законопроектът? (0 до 1) ").strip() or 0)

            effects["public_support"] = float(
                input("Колко обществено популярен е законопроектът? (-1 до 1) ").strip() or 0)

            effects["fiscal_impact"] = float(
                input("Какъв е бюджетният ефект на законопроекта? (-1 до 1) ").strip() or 0)

            effects["international_alignment"] = float(
                input("Доколко е съобразен с международните и европейските норми? (-1 до 1) ").strip() or 0)

            effects["urgency"] = float(
                input("Колко спешен е законопроектът? (0 до 1) ").strip() or 0)

            print("\nВъведи позициите на партиите между -1.0 и 1.0")
            for party_name in party_names:
                party_positions[party_name] = float(input(f"{party_name}: ").strip() or 0)

        custom_proposal = {
            "title": custom_title,
            "bill_id": custom_bill_id,
            "description": custom_description,
            "changes": custom_changes,
            "type": proposal_type,
            "candidate_party": candidate_party,
            "proposed_by_party": proposed_by_party,
            "effects": effects,
            "party_positions": party_positions
        }

        proposals.append(custom_proposal)

    print("\nЗаредени предложения:")
    for i, proposal in enumerate(proposals, start=1):
        print(f"{i}. {proposal['title']}")
        print(f"   ID: {proposal.get('bill_id', 'Няма ID')}")
        print(f"   За какво е: {proposal.get('description') or 'Няма добавено описание.'}")
        print(
            f"   Какво променя: {proposal.get('changes') or proposal.get('changes_summary') or 'Няма добавено описание на промените.'}")
        print(f"   Допълнителни характеристики: controversy={proposal.get('effects', {}).get('controversy', 0)}, public_support={proposal.get('effects', {}).get('public_support', 0)}, fiscal_impact={proposal.get('effects', {}).get('fiscal_impact', 0)}, international_alignment={proposal.get('effects', {}).get('international_alignment', 0)}, urgency={proposal.get('effects', {}).get('urgency', 0)}")
    choice = input("\nВъведи номер на предложение за гласуване или натисни Enter за всички: ").strip()

    if choice:
        selected_index = int(choice) - 1
        proposals_to_run = [proposals[selected_index]]
    else:
        proposals_to_run = proposals

    for proposal in proposals_to_run:
        auto_positions = calculate_party_positions(proposal)

        print("\nАвтоматично изчислени партийни позиции:")
        for party, position in auto_positions.items():
            print(f"{party}: {position}")

        simulation_output = get_representative_vote(
            bots, proposal, n_runs=200, pass_threshold=33
        )
        results = simulation_output["results"]
        totals = simulation_output["totals"]
        party_totals = simulation_output["party_totals"]
        bill_passed = simulation_output["bill_passed"]

        print(f"\nПредложение: {proposal['title']}")
        print(f"ID на законопроекта: {proposal.get('bill_id', 'Няма ID')}")
        description = proposal.get("description") or "Няма добавено описание на законопроекта."
        changes_text = proposal.get("changes") or proposal.get(
            "changes_summary"
        ) or "Няма добавено описание на промените."
        print(f"За какво е законопроектът: {description}")
        print(f"Какво променя: {changes_text}")
        print(
            f"Допълнителни характеристики: "
            f"controversy={proposal.get('effects', {}).get('controversy', 0)}, "
            f"public_support={proposal.get('effects', {}).get('public_support', 0)}, "
            f"fiscal_impact={proposal.get('effects', {}).get('fiscal_impact', 0)}, "
            f"international_alignment={proposal.get('effects', {}).get('international_alignment', 0)}, "
            f"urgency={proposal.get('effects', {}).get('urgency', 0)}\n"
        )

        for result in results:
            print(
                f"{result['name']} ({result['party']}): "
                f"{result['vote']} | резултат={result['score']} | "
                f"партия={result['party_pressure']} | идеология={result['ideology_score']} | "
                f"значимост={result['salience_score']} | отношение={result['relation_score']} | "
                f"шум={result['randomness']} | прагове=({result['yes_threshold']},{result['no_threshold']}) | "
                f"обяснение={result['reason']}"
            )

        print("\nОбщо:")
        print(f"ЗА: {totals['YES']}")
        print(f"ПРОТИВ: {totals['NO']}")
        print(f"ВЪЗДЪРЖАЛ СЕ: {totals['ABSTAIN']}")

        print("\nГласуване по партии:")
        for party, counts in party_totals.items():
            print(f"{party}: ЗА={counts['YES']} | ПРОТИВ={counts['NO']} | ВЪЗДЪРЖАЛ СЕ={counts['ABSTAIN']}")

        generate_pdf_report(proposal, results, totals, party_totals, bill_passed)
        print("\nPDF докладът е създаден: parliament_report.pdf")
